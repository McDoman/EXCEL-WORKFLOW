# -*- coding: utf-8 -*-
"""
Created on Wed Aug 27 18:45:51 2025

@author: LAdedo
"""

# pip install openpyxl
from pathlib import Path
from openpyxl import load_workbook, Workbook
import json

# ========== CONFIG (edit these) ==========
SOURCE_DIR = Path(r"C:/Users/ladedo/Desktop/MAINTENANCE FOLDER/2025/DT LIST/NEW PRIVATE DT FOLDER/martins 2")   # folder with your source Excel files
MASTER_PATH = Path(r"C:/Users/ladedo/Desktop/MAINTENANCE FOLDER/2025/DT LIST/NEW PRIVATE DT FOLDER/Master.xlsx")       # your holistic/master workbook
MASTER_SHEET_NAME = "Master"                        # target sheet in the master
INCLUDE_SUBDIRS = False                             # True to search subfolders
SOURCE_EXTENSIONS = {".xlsx", ".xlsm"}              # openpyxl doesn't read .xlsb/.xls
# =========================================

PROCESSED_LOG_PATH = MASTER_PATH.with_suffix(".processed.json")


def load_processed_log():
    if PROCESSED_LOG_PATH.exists():
        try:
            return json.loads(PROCESSED_LOG_PATH.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}
def save_processed_log(log):
    PROCESSED_LOG_PATH.write_text(json.dumps(log, indent=2), encoding="utf-8")

def load_or_create_master():
    if MASTER_PATH.exists():
        wb = load_workbook(MASTER_PATH)
    else:
        wb = Workbook()
    if MASTER_SHEET_NAME in wb.sheetnames:
        ws = wb[MASTER_SHEET_NAME]
    else:
        ws = wb.create_sheet(MASTER_SHEET_NAME)
        # Remove default empty "Sheet" if present
        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1 and wb["Sheet"]["A1"].value is None:
            wb.remove(wb["Sheet"])
    return wb, ws

def get_first_sheet(src_path):
    wb = load_workbook(src_path, data_only=True, read_only=True)
    ws = wb.worksheets[0]  # first sheet
    return wb, ws

def is_empty(ws):
    return ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None

def next_row(ws):
    return 1 if is_empty(ws) else ws.max_row + 1

def extract_values(ws):
    # Read required cells from the first sheet
    m6 = ws["M6"].value
    m7 = ws["M7"].value
    b21 = ws["B21"].value
    c5 = ws["C5"].value
    c8 = ws["C8"].value
    m8 = ws["M8"].value

    # Duplicate B21 into E and F
    e_val = b21
    f_val = b21

    return {
        "C": m6,
        "D": m7,
        "E": e_val,
        "F": f_val,
        "G": c5,
        "I": c8,
        "J": m8,
    }

def process_one_file(src_path, ws_master, processed_log):
    mtime = Path(src_path).stat().st_mtime
    key = str(Path(src_path).resolve())

    # Skip if unchanged since last run
    if key in processed_log and processed_log[key] == mtime:
        print(f"SKIP (unchanged): {src_path.name}")
        return False

    try:
        wb_src, ws_src = get_first_sheet(src_path)
        values = extract_values(ws_src)
        wb_src.close()
    except Exception as e:
        print(f"ERROR reading {src_path.name}: {e}")
        return False

    row = next_row(ws_master)
    for col_letter, val in values.items():
        ws_master[f"{col_letter}{row}"] = val

    processed_log[key] = mtime
    print(f"APPENDED row {row} from {src_path.name}")
    return True

def iter_source_files():
    if INCLUDE_SUBDIRS:
        files = (p for p in SOURCE_DIR.rglob("*") if p.is_file() and p.suffix.lower() in SOURCE_EXTENSIONS)
    else:
        files = (p for p in SOURCE_DIR.iterdir() if p.is_file() and p.suffix.lower() in SOURCE_EXTENSIONS)
    for p in files:
        if p.resolve() == MASTER_PATH.resolve():
            continue
        yield p

def main():
    if not SOURCE_DIR.exists():
        raise SystemExit(f"Source folder not found: {SOURCE_DIR}")

    processed_log = load_processed_log()
    wb_master, ws_master = load_or_create_master()

    changed = False
    for src in iter_source_files():
        if process_one_file(src, ws_master, processed_log):
            changed = True

    if changed:
        # Save master and log (close Excel if this fails on Windows)
        wb_master.save(MASTER_PATH)
        wb_master.close()
        save_processed_log(processed_log)
        print(f"\nSaved updates to {MASTER_PATH}")
    else:
        wb_master.close()
        print("\nNo new changes detected.")

if __name__ == "__main__":
    main()